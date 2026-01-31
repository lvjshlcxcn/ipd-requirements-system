"""Requirement service for business logic."""
from typing import Optional, List, Dict, Any, Tuple

from sqlalchemy.orm import Session

from app.models.requirement import Requirement, Requirement10QAnswer
from app.repositories.requirement import (
    RequirementRepository,
    Requirement10QRepository,
)
from app.repositories.workflow_history import WorkflowHistoryRepository
from app.schemas.requirement import (
    RequirementCreate,
    RequirementUpdate,
    Requirement10QCreate,
    RequirementResponse,
    RequirementStatsData,
)
from app.core.tenant import get_current_tenant


class RequirementService:
    """Service for Requirement business logic."""

    def __init__(self, db: Session):
        self.db = db
        self.repo = RequirementRepository(db)
        self.repo_10q = Requirement10QRepository(db)
        self.repo_history = WorkflowHistoryRepository(db)

    # ========================================================================
    # CRUD Operations
    # ========================================================================

    def create_requirement(
        self,
        data: RequirementCreate,
        created_by: Optional[int] = None,
    ) -> Requirement:
        """
        Create a new requirement with optional 10 questions.

        Args:
            data: Requirement creation data
            created_by: User ID who created the requirement

        Returns:
            Created requirement
        """
        # Get tenant_id from context
        tenant_id = get_current_tenant()

        # Convert 10q to dict if provided
        ten_q_data = None
        if data.customer_need_10q:
            ten_q_data = data.customer_need_10q.model_dump()

        # Create requirement
        requirement = self.repo.create(
            title=data.title,
            description=data.description,
            source_channel=data.source_channel,
            source_contact=data.source_contact,
            moscow_priority=data.moscow_priority,
            moscow_comment=data.moscow_comment,
            customer_need_10q=ten_q_data,
            estimated_duration_months=data.estimated_duration_months,
            complexity_level=data.complexity_level,
            created_by=created_by,
            tenant_id=tenant_id,
        )

        # Create detailed 10 questions answer if provided
        if data.customer_need_10q:
            self.repo_10q.create(
                requirement_id=requirement.id,
                answers=data.customer_need_10q.model_dump(),
                answered_by=created_by,
                tenant_id=tenant_id,
            )

        return requirement

    def get_requirement(self, requirement_id: int) -> Optional[Requirement]:
        """
        Get requirement by ID.

        Args:
            requirement_id: Requirement ID

        Returns:
            Requirement or None
        """
        return self.repo.get_by_id(requirement_id)

    def list_requirements(
        self,
        page: int = 1,
        page_size: int = 20,
        status: Optional[str] = None,
        source_channel: Optional[str] = None,
        target_type: Optional[str] = None,
        search: Optional[str] = None,
        sort_by: str = "created_at",
        sort_order: str = "desc",
    ) -> Tuple[List[Requirement], int]:
        """
        List requirements with filters and pagination.

        Args:
            page: Page number (1-indexed)
            page_size: Number of items per page
            status: Filter by status
            source_channel: Filter by source channel
            target_type: Filter by target type (sp/bp/charter/pcr)
            search: Search in title and requirement_no
            sort_by: Sort field
            sort_order: Sort order (asc or desc)

        Returns:
            Tuple of (list of requirements, total count)
        """
        return self.repo.list(
            page=page,
            page_size=page_size,
            status=status,
            source_channel=source_channel,
            target_type=target_type,
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
        )

    def update_requirement(
        self,
        requirement_id: int,
        data: RequirementUpdate,
        updated_by: Optional[int] = None,
    ) -> Optional[Requirement]:
        """
        Update requirement.

        Args:
            requirement_id: Requirement ID
            data: Update data
            updated_by: User ID who updated the requirement

        Returns:
            Updated requirement or None
        """
        requirement = self.repo.get_by_id(requirement_id)
        if not requirement:
            return None

        # Prepare updates
        updates = data.model_dump(exclude_unset=True, exclude={"customer_need_10q"})

        # Handle 10 questions update
        if data.customer_need_10q:
            ten_q_data = data.customer_need_10q.model_dump()
            updates["customer_need_10q"] = ten_q_data

            # Update or create detailed 10 questions answer
            existing_10q = self.repo_10q.get_by_requirement_id(requirement_id)
            if existing_10q:
                self.repo_10q.update(existing_10q, ten_q_data)
            else:
                # Use requirement's tenant_id for creating 10q
                self.repo_10q.create(
                    requirement_id=requirement_id,
                    answers=ten_q_data,
                    answered_by=updated_by,
                    tenant_id=requirement.tenant_id,
                )

        # Add updater info
        updates["updated_by"] = updated_by

        return self.repo.update(requirement, **updates)

    def delete_requirement(self, requirement_id: int) -> bool:
        """
        Delete requirement.

        Args:
            requirement_id: Requirement ID

        Returns:
            True if deleted, False if not found
        """
        requirement = self.repo.get_by_id(requirement_id)
        if not requirement:
            return False

        self.repo.delete(requirement)
        return True

    # ========================================================================
    # Status Operations
    # ========================================================================

    def update_status(
        self,
        requirement_id: int,
        new_status: str,
        updated_by: Optional[int] = None,
    ) -> Optional[Requirement]:
        """
        Update requirement status with automatic history tracking.

        Args:
            requirement_id: Requirement ID
            new_status: New status
            updated_by: User ID who updated the status

        Returns:
            Updated requirement or None
        """
        requirement = self.repo.get_by_id(requirement_id)
        if not requirement:
            return None

        # Record old status before updating
        old_status = requirement.status

        # Update status
        requirement = self.repo.update_status(requirement, new_status, updated_by)

        # Automatically record history
        self.repo_history.create(
            entity_type='requirement',
            entity_id=requirement_id,
            action='status_changed',
            from_status=old_status,
            to_status=new_status,
            performed_by=updated_by,
        )

        return requirement

    # ========================================================================
    # Statistics
    # ========================================================================

    def get_statistics(self) -> RequirementStatsData:
        """
        Get requirement statistics.

        Returns:
            Statistics data
        """
        total = self.repo.get_total_count()
        by_status = self.repo.get_stats_by_status()
        by_channel = self.repo.get_stats_by_channel()

        return RequirementStatsData(
            total=total,
            by_status=by_status,
            by_channel=by_channel,
        )

    # ========================================================================
    # 10 Questions
    # ========================================================================

    def get_10_questions(
        self, requirement_id: int
    ) -> Optional[Requirement10QAnswer]:
        """
        Get 10 questions answer by requirement ID.

        Args:
            requirement_id: Requirement ID

        Returns:
            10 questions answer or None
        """
        return self.repo_10q.get_by_requirement_id(requirement_id)

    # ========================================================================
    # History Tracking
    # ========================================================================

    def add_history_note(
        self,
        requirement_id: int,
        comments: str,
        action_reason: Optional[str] = None,
        performed_by: Optional[int] = None,
    ):
        """
        Add a manual note/comment to requirement history.

        Args:
            requirement_id: Requirement ID
            comments: Comment content
            action_reason: Optional reason for the note
            performed_by: User ID who added the note

        Returns:
            Created WorkflowHistory record
        """
        requirement = self.repo.get_by_id(requirement_id)
        if not requirement:
            raise ValueError("Requirement not found")

        return self.repo_history.create(
            entity_type='requirement',
            entity_id=requirement_id,
            action='note_added',
            to_status=requirement.status,
            comments=comments,
            action_reason=action_reason,
            performed_by=performed_by,
        )

    def get_history(
        self, requirement_id: int, limit: int = 50
    ) -> List:
        """
        Get requirement history records.

        Args:
            requirement_id: Requirement ID
            limit: Maximum number of records to return

        Returns:
            List of WorkflowHistory records
        """
        return self.repo_history.get_by_entity('requirement', requirement_id, limit)

    # ========================================================================
    # Export
    # ========================================================================

    @staticmethod
    def export_to_excel(
        db: Session,
        tenant_id: int,
        status: str = "distributed",
        target_type: str = "charter",
        search: Optional[str] = None
    ) -> bytes:
        """
        导出需求开发列表为 Excel.

        Args:
            db: 数据库会话
            tenant_id: 租户ID
            status: 需求状态（默认 distributed）
            target_type: 目标类型（默认 charter）
            search: 搜索关键词（可选）

        Returns:
            Excel 文件的字节流
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        import io

        # 获取需求数据
        repo = RequirementRepository(db)
        requirements, total = repo.list(
            page=1,
            page_size=10000,  # 大数量以支持导出所有数据
            status=status,
            target_type=target_type,
            search=search
        )

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "需求开发列表"

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 定义列宽
        column_widths = {
            'A': 18,  # 需求编号
            'B': 35,  # 需求标题
            'C': 12,  # 来源渠道
            'D': 10,  # 优先级
            'E': 12,  # MoSCoW
            'F': 15,  # Charter ID
            'G': 12,  # 预估工期
            'H': 18,  # 分发时间
            'I': 40,  # 需求描述
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 定义表头
        headers = [
            '需求编号',
            '需求标题',
            '来源渠道',
            '优先级',
            'MoSCoW优先级',
            'Charter ID',
            '预估工期',
            '分发时间',
            '需求描述'
        ]

        # 写入表头
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 来源渠道映射
        source_map = {
            'customer': '客户',
            'market': '市场',
            'sales': '销售',
            'rd': '技术',
            'after_sales': '售后',
            'competition': '竞争'
        }

        # 优先级映射
        priority_map = {
            90: '高',
            60: '中',
            30: '低'
        }

        # MoSCoW映射
        moscow_map = {
            'must_have': 'Must',
            'should_have': 'Should',
            'could_have': 'Could',
            'wont_have': "Won't"
        }

        # 写入数据
        for row_num, req in enumerate(requirements, 2):
            # 来源渠道
            source_text = source_map.get(req.source_channel, req.source_channel or '')

            # 优先级
            priority_text = priority_map.get(req.priority_score, '-') if req.priority_score else '-'

            # MoSCoW
            moscow_text = moscow_map.get(req.moscow_priority, '-') if req.moscow_priority else '-'

            # Charter ID
            charter_id = f"CHARTER-{str(req.target_id).zfill(3)}" if req.target_id else '-'

            # 预估工期
            duration_text = f"{req.estimated_duration_months}月" if req.estimated_duration_months else '-'

            # 分发时间
            updated_text = req.updated_at.strftime('%Y-%m-%d %H:%M') if req.updated_at else '-'

            # 需求描述（截取100字符）
            desc_text = (req.description[:100] + '...') if req.description and len(req.description) > 100 else (req.description or '-')

            ws.append([
                req.requirement_no or '',
                req.title or '',
                source_text,
                priority_text,
                moscow_text,
                charter_id,
                duration_text,
                updated_text,
                desc_text
            ])

            # 应用单元格样式
            for col_num in range(1, 10):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = cell_alignment
                cell.border = border

        # 添加汇总信息
        summary_row = len(requirements) + 3
        ws.cell(row=summary_row, column=1, value=f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
        ws.cell(row=summary_row, column=1).font = Font(name='微软雅黑', size=10)
        ws.cell(row=summary_row + 1, column=1, value=f'总计: {total} 条记录')
        ws.cell(row=summary_row + 1, column=1).font = Font(name='微软雅黑', size=10, bold=True)

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
