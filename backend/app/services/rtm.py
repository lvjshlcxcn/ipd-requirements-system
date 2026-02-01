"""RTM service for business logic."""
from typing import List, Optional, Dict, Any
from sqlalchemy.orm import Session, joinedload

from app.models.rtm import TraceabilityLink
from app.models.requirement import Requirement
from app.models.attachment import Attachment
from app.schemas.rtm import (
    TraceabilityLinkCreate,
    TraceabilityLinkUpdate,
    TraceabilityMatrix,
    TraceabilityItem,
    AttachmentInfo,
)


class RTMService:
    """需求追溯矩阵服务."""

    @staticmethod
    def get_traceability_matrix(
        db: Session,
        tenant_id: int,
        filters: Optional[Dict[str, Any]] = None,
    ) -> List[TraceabilityMatrix]:
        """
        获取需求追溯矩阵.

        Args:
            db: 数据库会话
            tenant_id: 租户ID
            filters: 过滤条件

        Returns:
            追溯矩阵列表
        """
        # 获取租户下的所有需求
        requirements_query = db.query(Requirement).filter(
            Requirement.tenant_id == tenant_id
        )

        requirements = requirements_query.all()

        # 构建追溯矩阵
        matrix = []
        for req in requirements:
            # 获取该需求的所有追溯关联（预加载附件信息）
            links = db.query(TraceabilityLink).options(
                joinedload(TraceabilityLink.design_attachment),
                joinedload(TraceabilityLink.code_attachment),
                joinedload(TraceabilityLink.test_attachment)
            ).filter(
                TraceabilityLink.requirement_id == req.id,
                TraceabilityLink.tenant_id == tenant_id,
            ).all()

            # 分类关联项
            design_items = []
            code_items = []
            test_items = []

            for link in links:
                # 构建附件信息
                design_att = AttachmentInfo.model_validate(link.design_attachment) if link.design_attachment else None
                code_att = AttachmentInfo.model_validate(link.code_attachment) if link.code_attachment else None
                test_att = AttachmentInfo.model_validate(link.test_attachment) if link.test_attachment else None

                # 根据关联类型分类
                if link.design_id or link.design_attachment_id:
                    design_items.append(
                        TraceabilityItem(
                            id=link.id,
                            design_id=link.design_id,
                            design_attachment_id=link.design_attachment_id,
                            design_attachment=design_att
                        )
                    )
                if link.code_id or link.code_attachment_id:
                    code_items.append(
                        TraceabilityItem(
                            id=link.id,
                            code_id=link.code_id,
                            code_attachment_id=link.code_attachment_id,
                            code_attachment=code_att
                        )
                    )
                if link.test_id or link.test_attachment_id:
                    test_items.append(
                        TraceabilityItem(
                            id=link.id,
                            test_id=link.test_id,
                            test_attachment_id=link.test_attachment_id,
                            test_attachment=test_att
                        )
                    )

            matrix.append(
                TraceabilityMatrix(
                    requirement_id=req.id,  # 数据库ID
                    requirement_no=req.requirement_no,  # 使用业务需求编号
                    requirement_title=req.title or f"需求 {req.requirement_no}",
                    design_items=design_items,
                    code_items=code_items,
                    test_items=test_items,
                )
            )

        return matrix

    @staticmethod
    def get_requirement_traceability(
        db: Session,
        requirement_id: int,
        tenant_id: int,
    ) -> Optional[TraceabilityMatrix]:
        """
        获取单个需求的追溯信息.

        Args:
            db: 数据库会话
            requirement_id: 需求ID
            tenant_id: 租户ID

        Returns:
            追溯矩阵
        """
        req = db.query(Requirement).filter(
            Requirement.id == requirement_id,
            Requirement.tenant_id == tenant_id,
        ).first()

        if not req:
            return None

        links = db.query(TraceabilityLink).options(
            joinedload(TraceabilityLink.design_attachment),
            joinedload(TraceabilityLink.code_attachment),
            joinedload(TraceabilityLink.test_attachment)
        ).filter(
            TraceabilityLink.requirement_id == requirement_id,
            TraceabilityLink.tenant_id == tenant_id,
        ).all()

        design_items = []
        code_items = []
        test_items = []

        for link in links:
            design_att = AttachmentInfo.model_validate(link.design_attachment) if link.design_attachment else None
            code_att = AttachmentInfo.model_validate(link.code_attachment) if link.code_attachment else None
            test_att = AttachmentInfo.model_validate(link.test_attachment) if link.test_attachment else None

            if link.design_id or link.design_attachment_id:
                design_items.append(
                    TraceabilityItem(
                        id=link.id,
                        design_id=link.design_id,
                        design_attachment_id=link.design_attachment_id,
                        design_attachment=design_att
                    )
                )
            if link.code_id or link.code_attachment_id:
                code_items.append(
                    TraceabilityItem(
                        id=link.id,
                        code_id=link.code_id,
                        code_attachment_id=link.code_attachment_id,
                        code_attachment=code_att
                    )
                )
            if link.test_id or link.test_attachment_id:
                test_items.append(
                    TraceabilityItem(
                        id=link.id,
                        test_id=link.test_id,
                        test_attachment_id=link.test_attachment_id,
                        test_attachment=test_att
                    )
                )

        return TraceabilityMatrix(
            requirement_id=req.id,  # 数据库ID
            requirement_no=req.requirement_no,  # 使用业务需求编号
            requirement_title=req.title or f"需求 {req.requirement_no}",
            design_items=design_items,
            code_items=code_items,
            test_items=test_items,
        )

    @staticmethod
    def create_link(
        db: Session,
        link_data: TraceabilityLinkCreate,
        tenant_id: int,
    ) -> TraceabilityLink:
        """
        创建追溯关联.

        Args:
            db: 数据库会话
            link_data: 关联数据
            tenant_id: 租户ID

        Returns:
            创建的关联
        """
        # 验证需求存在
        req = db.query(Requirement).filter(
            Requirement.id == link_data.requirement_id,
            Requirement.tenant_id == tenant_id,
        ).first()

        if not req:
            raise ValueError(f"需求 {link_data.requirement_id} 不存在")

        # 至少需要一个追溯项（文档ID或附件ID）
        has_design = link_data.design_id or link_data.design_attachment_id
        has_code = link_data.code_id or link_data.code_attachment_id
        has_test = link_data.test_id or link_data.test_attachment_id

        if not any([has_design, has_code, has_test]):
            raise ValueError("至少需要提供一个追溯项（设计文档/代码/测试用例）")

        # 如果提供了附件ID，验证附件存在
        if link_data.design_attachment_id:
            att = db.query(Attachment).filter(
                Attachment.id == link_data.design_attachment_id
            ).first()
            if not att:
                raise ValueError(f"设计文档附件 {link_data.design_attachment_id} 不存在")

        if link_data.code_attachment_id:
            att = db.query(Attachment).filter(
                Attachment.id == link_data.code_attachment_id
            ).first()
            if not att:
                raise ValueError(f"代码附件 {link_data.code_attachment_id} 不存在")

        if link_data.test_attachment_id:
            att = db.query(Attachment).filter(
                Attachment.id == link_data.test_attachment_id
            ).first()
            if not att:
                raise ValueError(f"测试用例附件 {link_data.test_attachment_id} 不存在")

        # 创建关联
        link = TraceabilityLink(
            requirement_id=link_data.requirement_id,
            design_id=link_data.design_id,
            code_id=link_data.code_id,
            test_id=link_data.test_id,
            design_attachment_id=link_data.design_attachment_id,
            code_attachment_id=link_data.code_attachment_id,
            test_attachment_id=link_data.test_attachment_id,
            notes=link_data.notes,
            tenant_id=tenant_id,
        )

        db.add(link)
        db.commit()
        db.refresh(link)

        return link

    @staticmethod
    def update_link(
        db: Session,
        link_id: int,
        link_data: TraceabilityLinkUpdate,
        tenant_id: int,
    ) -> Optional[TraceabilityLink]:
        """
        更新追溯关联.

        Args:
            db: 数据库会话
            link_id: 关联ID
            link_data: 更新数据
            tenant_id: 租户ID

        Returns:
            更新后的关联
        """
        link = db.query(TraceabilityLink).filter(
            TraceabilityLink.id == link_id,
            TraceabilityLink.tenant_id == tenant_id,
        ).first()

        if not link:
            return None

        # 更新字段
        if link_data.design_id is not None:
            link.design_id = link_data.design_id
        if link_data.code_id is not None:
            link.code_id = link_data.code_id
        if link_data.test_id is not None:
            link.test_id = link_data.test_id

        # 更新附件ID字段
        if link_data.design_attachment_id is not None:
            # 验证附件存在
            if link_data.design_attachment_id:
                att = db.query(Attachment).filter(
                    Attachment.id == link_data.design_attachment_id
                ).first()
                if not att:
                    raise ValueError(f"设计文档附件 {link_data.design_attachment_id} 不存在")
            link.design_attachment_id = link_data.design_attachment_id

        if link_data.code_attachment_id is not None:
            if link_data.code_attachment_id:
                att = db.query(Attachment).filter(
                    Attachment.id == link_data.code_attachment_id
                ).first()
                if not att:
                    raise ValueError(f"代码附件 {link_data.code_attachment_id} 不存在")
            link.code_attachment_id = link_data.code_attachment_id

        if link_data.test_attachment_id is not None:
            if link_data.test_attachment_id:
                att = db.query(Attachment).filter(
                    Attachment.id == link_data.test_attachment_id
                ).first()
                if not att:
                    raise ValueError(f"测试用例附件 {link_data.test_attachment_id} 不存在")
            link.test_attachment_id = link_data.test_attachment_id

        if link_data.notes is not None:
            link.notes = link_data.notes
        if link_data.status is not None:
            link.status = link_data.status

        db.commit()
        db.refresh(link)

        return link

    @staticmethod
    def delete_link(
        db: Session,
        link_id: int,
        tenant_id: int,
    ) -> bool:
        """
        删除追溯关联.

        Args:
            db: 数据库会话
            link_id: 关联ID
            tenant_id: 租户ID

        Returns:
            是否删除成功
        """
        link = db.query(TraceabilityLink).filter(
            TraceabilityLink.id == link_id,
            TraceabilityLink.tenant_id == tenant_id,
        ).first()

        if not link:
            return False

        db.delete(link)
        db.commit()

        return True

    @staticmethod
    def export_to_excel(
        db: Session,
        tenant_id: int,
    ) -> bytes:
        """
        导出需求追溯矩阵为 Excel.

        Args:
            db: 数据库会话
            tenant_id: 租户ID

        Returns:
            Excel 文件的字节流
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        import io

        # 获取追溯矩阵数据
        matrix = rtm_service.get_traceability_matrix(db, tenant_id)

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "需求追溯矩阵"

        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 设置列宽
        ws.column_dimensions['A'].width = 15  # 需求ID
        ws.column_dimensions['B'].width = 40  # 需求标题
        ws.column_dimensions['C'].width = 30  # 设计文档
        ws.column_dimensions['D'].width = 30  # 代码
        ws.column_dimensions['E'].width = 30  # 测试用例
        ws.column_dimensions['F'].width = 12  # 状态

        # 写入表头
        headers = ['需求ID', '需求标题', '设计文档', '代码', '测试用例', '追溯状态']
        ws.append(headers)

        # 应用表头样式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 写入数据
        for row_num, item in enumerate(matrix, 2):
            # 获取追溯项
            design_items = ', '.join([d.design_id or '' for d in item.design_items]) if item.design_items else '无'
            code_items = ', '.join([c.code_id or '' for c in item.code_items]) if item.code_items else '无'
            test_items = ', '.join([t.test_id or '' for t in item.test_items]) if item.test_items else '无'

            # 计算状态
            has_design = len(item.design_items) > 0
            has_code = len(item.code_items) > 0
            has_test = len(item.test_items) > 0

            if has_design and has_code and has_test:
                status = '完整'
            elif has_design or has_code or has_test:
                status = '部分'
            else:
                status = '缺失'

            ws.append([
                item.requirement_id,
                item.requirement_title,
                design_items,
                code_items,
                test_items,
                status
            ])

            # 应用单元格样式
            for col_num in range(1, 7):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = cell_alignment
                cell.border = border

        # 添加汇总信息
        row_num = len(matrix) + 3
        ws.cell(row=row_num, column=1, value='汇总统计')
        ws.cell(row=row_num, column=1).font = Font(name='微软雅黑', size=12, bold=True)
        row_num += 1

        total = len(matrix)
        complete = sum(1 for item in matrix
                      if len(item.design_items) > 0 and len(item.code_items) > 0 and len(item.test_items) > 0)
        partial = sum(1 for item in matrix
                     if (len(item.design_items) > 0 or len(item.code_items) > 0 or len(item.test_items) > 0)
                     and not (len(item.design_items) > 0 and len(item.code_items) > 0 and len(item.test_items) > 0))
        missing = sum(1 for item in matrix if len(item.design_items) == 0 and len(item.code_items) == 0 and len(item.test_items) == 0)

        ws.cell(row=row_num, column=1, value=f'总计需求: {total}')
        row_num += 1
        ws.cell(row=row_num, column=1, value=f'完整追溯: {complete}')
        row_num += 1
        ws.cell(row=row_num, column=1, value=f'部分追溯: {partial}')
        row_num += 1
        ws.cell(row=row_num, column=1, value=f'缺失追溯: {missing}')
        row_num += 2
        ws.cell(row=row_num, column=1, value=f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')

        # 保存到字节流
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read()

    @staticmethod
    def get_statistics(
        db: Session,
        tenant_id: int,
    ) -> dict:
        """
        获取追溯矩阵统计数据.

        Args:
            db: 数据库会话
            tenant_id: 租户ID

        Returns:
            统计数据字典
        """
        matrix = rtm_service.get_traceability_matrix(db, tenant_id)

        total = len(matrix)
        complete = sum(1 for item in matrix
                      if len(item.design_items) > 0 and len(item.code_items) > 0 and len(item.test_items) > 0)
        partial = sum(1 for item in matrix
                     if (len(item.design_items) > 0 or len(item.code_items) > 0 or len(item.test_items) > 0)
                     and not (len(item.design_items) > 0 and len(item.code_items) > 0 and len(item.test_items) > 0))
        missing = sum(1 for item in matrix if len(item.design_items) == 0 and len(item.code_items) == 0 and len(item.test_items) == 0)

        # 统计设计、代码、测试的覆盖率
        with_design = sum(1 for item in matrix if len(item.design_items) > 0)
        with_code = sum(1 for item in matrix if len(item.code_items) > 0)
        with_test = sum(1 for item in matrix if len(item.test_items) > 0)

        return {
            'total': total,
            'complete': complete,
            'partial': partial,
            'missing': missing,
            'coverage': {
                'design': with_design,
                'code': with_code,
                'test': with_test,
            },
            'completion_rate': round((complete / total * 100) if total > 0 else 0, 2),
        }


rtm_service = RTMService()
