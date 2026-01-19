"""Excel import/export utility."""
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


class ExcelHandler:
    """Handler for Excel import/export operations."""

    @staticmethod
    def create_template() -> bytes:
        """
        Create an Excel template for requirement import.

        Returns:
            Excel file as bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requirements"

        # Define headers
        headers = [
            "需求标题",
            "需求描述",
            "用户角色",
            "用户行动",
            "用户收益",
            "优先级分数",
            "MoSCoW优先级",
            "Kano分类",
            "业务价值",
            "技术复杂度",
            "预计工作量",
            "来源",
            "标签",
        ]

        # Set header styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Add dropdown instructions
        ws.cell(row=2, column=1, value="")
        ws.cell(row=2, column=2, value="填写说明：")
        ws.cell(row=3, column=2, value="1. MoSCoW优先级: must_have(必须有), should_have(应该有), could_have(可以有), wont_have(本次不做)")
        ws.cell(row=4, column=2, value="2. Kano分类: excitement(兴奋型), performance(期望型), basic(必备型), indifferent(无差异型), reverse(反向型)")
        ws.cell(row=5, column=2, value="3. 优先级分数: 1-10的数字")
        ws.cell(row=6, column=2, value="4. 预计工作量: 小时数")

        # Adjust column widths
        column_widths = {
            1: 30,  # 需求标题
            2: 50,  # 需求描述
            3: 15,  # 用户角色
            4: 30,  # 用户行动
            5: 30,  # 用户收益
            6: 12,  # 优先级分数
            7: 15,  # MoSCoW优先级
            8: 15,  # Kano分类
            9: 12,  # 业务价值
            10: 12,  # 技术复杂度
            11: 12,  # 预计工作量
            12: 20,  # 来源
            13: 30,  # 标签
        }

        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def import_from_excel(file_path: str, skip_header: bool = True) -> List[Dict[str, Any]]:
        """
        Import requirements from Excel file.

        Args:
            file_path: Path to Excel file
            skip_header: Whether to skip header row

        Returns:
            List of requirement dictionaries
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        requirements = []
        start_row = 2 if skip_header else 1

        for row_num in range(start_row, ws.max_row + 1):
            # Get values from row
            values = [ws.cell(row_num, col).value for col in range(1, ws.max_column + 1)]

            # Skip empty rows
            if not any(values):
                continue

            # Build requirement dict
            req = {
                "title": values[0] or "",
                "description": values[1] or "",
                "user_role": values[2] or "",
                "user_action": values[3] or "",
                "user_benefit": values[4] or "",
                "priority_score": float(values[5]) if values[5] else 5.0,
                "moscow_priority": values[6] or "should_have",
                "kano_category": values[7] or "performance",
                "business_value": int(values[8]) if values[8] else 5,
                "technical_complexity": int(values[9]) if values[9] else 5,
                "estimated_hours": int(values[10]) if values[10] else 0,
                "source": values[11] or "",
                "tags": values[12] or "",
            }

            requirements.append(req)

        return requirements

    @staticmethod
    def export_to_excel(
        data: List[Dict[str, Any]], sheet_name: str = "Requirements"
    ) -> bytes:
        """
        Export data to Excel file.

        Args:
            data: List of dictionaries to export
            sheet_name: Name of the sheet

        Returns:
            Excel file as bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            return io.BytesIO().getvalue()

        # Get headers from first row
        headers = list(data[0].keys())

        # Set header styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Write data
        for row_num, item in enumerate(data, start=2):
            for col_num, key in enumerate(headers, start=1):
                value = item.get(key, "")

                # Format datetime values
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                # Format dict values
                elif isinstance(value, (dict, list)):
                    value = str(value)

                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col_num)

            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    @staticmethod
    def export_verification_checklist(
        checklists: List[Dict[str, Any]], requirement_title: str = ""
    ) -> bytes:
        """
        Export verification checklists to Excel.

        Args:
            checklists: List of checklist dictionaries
            requirement_title: Title of the requirement

        Returns:
            Excel file as bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Verification Checklists"

        # Title
        ws.cell(row=1, column=1, value=f"验证检查清单 - {requirement_title}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Headers
        headers = ["验证类型", "检查项", "状态", "备注"]

        # Set header styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        row_num = 4
        for checklist in checklists:
            verification_type = checklist.get("verification_type", "")
            checklist_items = checklist.get("checklist_items", [])

            for item in checklist_items:
                ws.cell(row=row_num, column=1, value=verification_type)
                ws.cell(row=row_num, column=2, value=item.get("item", ""))
                ws.cell(row=row_num, column=3, value="已完成" if item.get("checked") else "未完成")
                ws.cell(row=row_num, column=4, value=item.get("notes", ""))
                row_num += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 30

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
